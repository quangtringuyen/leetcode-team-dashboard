"""
Team management endpoints
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from backend.core.security import get_current_user
from backend.core.storage import read_json, write_json
from backend.core.config import settings
from backend.utils.leetcodeapi import fetch_user_data

router = APIRouter()

def get_members_list_internal(username: str) -> List[Dict[str, Any]]:
    """Helper to get members list for internal use"""
    all_members = read_json(settings.MEMBERS_FILE, default={})
    return all_members.get(username, [])

class TeamMember(BaseModel):
    username: str
    name: Optional[str] = None
    status: Optional[str] = "active"  # active, suspended

class TeamMemberResponse(BaseModel):
    username: str
    name: str
    status: str = "active"
    avatar: Optional[str] = None
    totalSolved: int = 0
    ranking: Optional[int] = None
    easy: int = 0
    medium: int = 0
    hard: int = 0

@router.get("/members", response_model=List[TeamMemberResponse])
async def get_team_members(current_user: dict = Depends(get_current_user)):
    """Get all team members for current user"""
    username = current_user["username"]

    # Load all members
    all_members = read_json(settings.MEMBERS_FILE, default={})
    user_members = all_members.get(username, [])

    # Fetch fresh data for each member - PARALLELIZED
    result = []

    def fetch_member_data(member):
        """Helper function to fetch data for a single member"""
        try:
            leetcode_data = fetch_user_data(member["username"])
            if leetcode_data:
                return TeamMemberResponse(
                    username=member["username"],
                    name=member.get("name", member["username"]),
                    status=member.get("status", "active"),
                    avatar=leetcode_data.get("avatar"),
                    totalSolved=leetcode_data.get("totalSolved", 0),
                    ranking=leetcode_data.get("ranking"),
                    easy=leetcode_data.get("easy", 0),
                    medium=leetcode_data.get("medium", 0),
                    hard=leetcode_data.get("hard", 0)
                )
            else:
                # If fetch fails, return basic data
                return TeamMemberResponse(
                    username=member["username"],
                    name=member.get("name", member["username"]),
                    status=member.get("status", "active"),
                    totalSolved=0,
                    easy=0,
                    medium=0,
                    hard=0
                )
        except Exception as e:
            # If error, return basic data
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error fetching data for {member['username']}: {e}")
            return TeamMemberResponse(
                username=member["username"],
                name=member.get("name", member["username"]),
                status=member.get("status", "active"),
                totalSolved=0,
                easy=0,
                medium=0,
                hard=0
            )

    # Fetch all member data in parallel
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(fetch_member_data, member) for member in user_members]

        for future in as_completed(futures):
            try:
                member_response = future.result()
                result.append(member_response)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error processing member data: {e}")

    # Sort by totalSolved descending
    result.sort(key=lambda x: x.totalSolved, reverse=True)

    return result

@router.post("/members", status_code=status.HTTP_201_CREATED)
async def add_team_member(
    member: TeamMember,
    current_user: dict = Depends(get_current_user)
):
    """Add a new team member"""
    username = current_user["username"]

    # Verify LeetCode user exists
    leetcode_data = fetch_user_data(member.username)
    if not leetcode_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"LeetCode user '{member.username}' not found"
        )

    # Load all members
    all_members = read_json(settings.MEMBERS_FILE, default={})
    user_members = all_members.get(username, [])

    # Check if already exists
    if any(m["username"] == member.username for m in user_members):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"User '{member.username}' is already in your team"
        )

    # Add new member
    user_members.append({
        "username": member.username,
        "name": member.name or leetcode_data.get("realName", member.username),
        "status": member.status or "active"
    })

    all_members[username] = user_members
    write_json(settings.MEMBERS_FILE, all_members)

    return {
        "message": f"Added {member.username} to team",
        "member": {
            "username": member.username,
            "name": member.name or leetcode_data.get("realName", member.username),
            "status": member.status or "active"
        }
    }

class MemberUpdate(BaseModel):
    name: Optional[str] = None
    username: Optional[str] = None
    status: Optional[str] = None  # active, suspended

@router.put("/members/{member_username}")
async def update_team_member(
    member_username: str,
    update_data: MemberUpdate,
    current_user: dict = Depends(get_current_user)
):
    """
    Update a team member's details.
    Allows changing name (display name) and username (LeetCode handle).
    If name is blank, it defaults to username.
    """
    username = current_user["username"]
    
    # Load all members
    all_members = read_json(settings.MEMBERS_FILE, default={})
    user_members = all_members.get(username, [])
    
    # Find member
    member_idx = -1
    member = None
    for i, m in enumerate(user_members):
        if m["username"].lower() == member_username.lower():
            member_idx = i
            member = m
            break
            
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
        
    # Update Name
    if update_data.name is not None:
        # If name is blank/empty string, set it to None (which will fallback to username in display)
        # But user requirement: "if name is blank, display username"
        # In our system, we usually store `name` field.
        # If we store it as empty string, frontend should handle it.
        # Or we can store it as the username.
        if update_data.name.strip() == "":
            member["name"] = update_data.username if update_data.username else member_username
        else:
            member["name"] = update_data.name
            
    # Update Username (Complex operation)
    if update_data.username and update_data.username != member_username:
        new_username = update_data.username
        
        # Check if new username exists on LeetCode
        leetcode_data = fetch_user_data(new_username)
        if not leetcode_data:
            raise HTTPException(status_code=404, detail=f"LeetCode user '{new_username}' not found")
            
        # Check if already in team
        if any(m["username"] == new_username for m in user_members):
            raise HTTPException(status_code=400, detail=f"User '{new_username}' is already in your team")
            
        # Update in JSON
        member["username"] = new_username
        
        # If name was defaulting to old username, update it to new username
        if member.get("name") == member_username:
            member["name"] = new_username
            
        # Update in Database (History/Snapshots)
        # We need to update all snapshots associated with old username to new username
        from backend.core.database import get_db_connection
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Update members table if exists
                # Check if we use members table in DB (yes we do)
                # But we can't update PK easily.
                # Strategy: 
                # 1. Create new member in DB
                # 2. Update snapshots to point to new member
                # 3. Delete old member
                
                # 1. Insert new member (ignore if exists)
                cursor.execute("""
                INSERT OR IGNORE INTO members (username, name, team_owner)
                SELECT ?, ?, team_owner FROM members WHERE username = ?
                """, (new_username, member.get("name"), member_username))
                
                # 2. Update snapshots
                cursor.execute("""
                UPDATE snapshots SET username = ? WHERE username = ?
                """, (new_username, member_username))
                
                # 3. Delete old member
                cursor.execute("DELETE FROM members WHERE username = ?", (member_username,))
                
                conn.commit()
        except Exception as e:
            # Log error but continue since JSON is primary for members list currently
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error updating database for member rename: {e}")

            logger = logging.getLogger(__name__)
            logger.error(f"Error updating database for member rename: {e}")

    # Update Status
    if update_data.status:
        member["status"] = update_data.status

    # Save changes
    user_members[member_idx] = member
    all_members[username] = user_members
    write_json(settings.MEMBERS_FILE, all_members)
    
    return {"message": "Member updated successfully", "member": member}

@router.delete("/members/{member_username}")
async def remove_team_member(
    member_username: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove a team member"""
    username = current_user["username"]

    # Load all members
    all_members = read_json(settings.MEMBERS_FILE, default={})
    user_members = all_members.get(username, [])

    # Find and remove member
    user_members = [m for m in user_members if m["username"] != member_username]

    all_members[username] = user_members
    write_json(settings.MEMBERS_FILE, all_members)

    return {"message": f"Removed {member_username} from team"}

@router.get("/stats")
async def get_team_stats(current_user: dict = Depends(get_current_user)):
    """Get overall team statistics"""
    username = current_user["username"]

    # Load members
    all_members = read_json(settings.MEMBERS_FILE, default={})
    user_members_raw = all_members.get(username, [])
    # Filter out suspended members
    user_members = [m for m in user_members_raw if m.get("status", "active") != "suspended"]

    if not user_members:
        return {
            "total_members": 0,
            "total_problems_solved": 0,
            "average_solved": 0,
            "difficulty_breakdown": {
                "easy": 0,
                "medium": 0,
                "hard": 0
            }
        }

    # Fetch data for all members - PARALLELIZED
    members_data = []

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [executor.submit(fetch_user_data, member["username"]) for member in user_members]

        for future in as_completed(futures):
            try:
                data = future.result()
                if data:
                    members_data.append(data)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error fetching member data: {e}")

    total_solved = sum(m.get("totalSolved", 0) for m in members_data)
    avg_solved = total_solved // len(members_data) if members_data else 0

    # Aggregate difficulty breakdown
    total_easy = sum(m.get("easy", 0) for m in members_data)
    total_medium = sum(m.get("medium", 0) for m in members_data)
    total_hard = sum(m.get("hard", 0) for m in members_data)

    return {
        "total_members": len(user_members),
        "total_problems_solved": total_solved,
        "average_solved": avg_solved,
        "difficulty_breakdown": {
            "easy": total_easy,
            "medium": total_medium,
            "hard": total_hard
        }
    }

@router.get("/export/excel")
async def export_team_excel(current_user: dict = Depends(get_current_user)):
    """Export team data to Excel format"""
    username = current_user["username"]
    
    # Get team members with fresh data
    all_members = read_json(settings.MEMBERS_FILE, default={})
    user_members = all_members.get(username, [])
    
    # Fetch fresh data for each member
    members_data = []
    for member in user_members:
        leetcode_data = fetch_user_data(member["username"])
        if leetcode_data:
            members_data.append({
                "username": member["username"],
                "name": member.get("name", member["username"]),
                "totalSolved": leetcode_data.get("totalSolved", 0),
                "easy": leetcode_data.get("easy", 0),
                "medium": leetcode_data.get("medium", 0),
                "hard": leetcode_data.get("hard", 0),
                "ranking": leetcode_data.get("ranking", "N/A")
            })
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Team Members Sheet
    ws_members = wb.active
    ws_members.title = "Team Members"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Username", "Name", "Total Solved", "Easy", "Medium", "Hard", "Ranking"]
    for col, header in enumerate(headers, 1):
        cell = ws_members.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row, member in enumerate(members_data, 2):
        ws_members.cell(row=row, column=1, value=member["username"])
        ws_members.cell(row=row, column=2, value=member["name"])
        ws_members.cell(row=row, column=3, value=member["totalSolved"])
        ws_members.cell(row=row, column=4, value=member.get("easy", 0))
        ws_members.cell(row=row, column=5, value=member.get("medium", 0))
        ws_members.cell(row=row, column=6, value=member.get("hard", 0))
        ws_members.cell(row=row, column=7, value=str(member["ranking"]))
    
    # Auto-adjust column widths
    for column in ws_members.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_members.column_dimensions[column_letter].width = max_length + 2
    
    # Statistics Sheet
    ws_stats = wb.create_sheet("Statistics")
    ws_stats.cell(row=1, column=1, value="Metric").font = header_font
    ws_stats.cell(row=1, column=2, value="Value").font = header_font
    
    total_solved = sum(m["totalSolved"] for m in members_data)
    total_easy = sum(m.get("easy", 0) for m in members_data)
    total_medium = sum(m.get("medium", 0) for m in members_data)
    total_hard = sum(m.get("hard", 0) for m in members_data)
    avg_solved = total_solved // len(members_data) if members_data else 0
    
    stats = [
        ("Total Members", len(members_data)),
        ("Total Problems Solved", total_solved),
        ("Average Solved per Member", avg_solved),
        ("Total Easy", total_easy),
        ("Total Medium", total_medium),
        ("Total Hard", total_hard)
    ]
    
    for row, (metric, value) in enumerate(stats, 2):
        ws_stats.cell(row=row, column=1, value=metric)
        ws_stats.cell(row=row, column=2, value=value)
    
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 15
    
    # Week-over-Week Analytics Sheet
    ws_wow = wb.create_sheet("Week-over-Week")
    ws_wow.title = "Week-over-Week"

    # Headers for Week-over-Week
    wow_headers = ["Week", "Member", "Previous", "Current", "Change", "% Change", "Rank", "Rank Δ"]
    for col, header in enumerate(wow_headers, 1):
        cell = ws_wow.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Fetch week-over-week data from analytics module
    from backend.api.analytics import get_week_over_week_internal
    
    # Get all available historical data (52 weeks = 1 year)
    wow_data = get_week_over_week_internal(username, weeks=52)
    
    # Data rows for Week-over-Week
    for row_idx, entry in enumerate(wow_data, 2):
        ws_wow.cell(row=row_idx, column=1, value=entry.get("week", ""))
        ws_wow.cell(row=row_idx, column=2, value=entry.get("member", ""))
        ws_wow.cell(row=row_idx, column=3, value=entry.get("previous", 0))
        ws_wow.cell(row=row_idx, column=4, value=entry.get("current", 0))
        ws_wow.cell(row=row_idx, column=5, value=entry.get("change", 0))
        ws_wow.cell(row=row_idx, column=6, value=f"{entry.get('pct_change', 0)}%")
        ws_wow.cell(row=row_idx, column=7, value=entry.get("rank", 0))
        
        # Format rank delta with + or - sign
        rank_delta = entry.get("rank_delta", 0)
        if rank_delta > 0:
            ws_wow.cell(row=row_idx, column=8, value=f"+{rank_delta}")
        elif rank_delta < 0:
            ws_wow.cell(row=row_idx, column=8, value=str(rank_delta))
        else:
            ws_wow.cell(row=row_idx, column=8, value="0")

    # Auto-adjust column widths for Week-over-Week sheet
    for column in ws_wow.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_wow.column_dimensions[column_letter].width = max_length + 2
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with current date
    filename = f"team-data-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/backup")
async def backup_team_data(current_user: dict = Depends(get_current_user)):
    """Export all user data as JSON backup"""
    username = current_user["username"]
    
    # Load all data files
    all_members = read_json(settings.MEMBERS_FILE, default={})
    all_history = read_json(settings.HISTORY_FILE, default={})
    all_snapshots = read_json(settings.SNAPSHOTS_FILE, default={})
    
    # Extract user-specific data
    backup_data = {
        "version": "1.0",
        "exported_at": datetime.now().isoformat(),
        "username": username,
        "members": all_members.get(username, []),
        "history": all_history.get(username, []),
        "snapshots": all_snapshots.get(username, [])
    }
    
    return backup_data

class RestoreBackup(BaseModel):
    version: str
    username: str
    members: List[Dict[str, Any]]
    history: Optional[List[Dict[str, Any]]] = []
    snapshots: Optional[List[Dict[str, Any]]] = []

@router.post("/restore")
async def restore_team_data(
    backup: RestoreBackup,
    current_user: dict = Depends(get_current_user)
):
    """Restore data from JSON backup"""
    username = current_user["username"]
    
    # Validate backup version
    if backup.version != "1.0":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported backup version: {backup.version}"
        )
    
    # Load current data
    all_members = read_json(settings.MEMBERS_FILE, default={})
    all_history = read_json(settings.HISTORY_FILE, default={})
    all_snapshots = read_json(settings.SNAPSHOTS_FILE, default={})
    
    # Restore data
    all_members[username] = backup.members
    all_history[username] = backup.history or []
    all_snapshots[username] = backup.snapshots or []
    
    # Write back to files
    write_json(settings.MEMBERS_FILE, all_members)
    write_json(settings.HISTORY_FILE, all_history)
    write_json(settings.SNAPSHOTS_FILE, all_snapshots)
    
    return {
        "message": "Data restored successfully",
        "members_restored": len(backup.members),
        "history_entries": len(backup.history or []),
        "snapshots_restored": len(backup.snapshots or [])
    }
